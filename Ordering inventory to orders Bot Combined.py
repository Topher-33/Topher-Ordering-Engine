from __future__ import print_function
import csv
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl import Workbook
import xlsxwriter
import os.path
import tkinter as tk
from tkinter import messagebox
import tkinter
from tkinter import *


#########################################################
#########################################################
#            CHANGE THIS FOR NEW CLIENT                 #
#########################################################
##############################################################################################
""" THIS LINE CHOOSES WHAT GOOGLE SPREAD SHEET TO ACCESS"""
scope = ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]
""" THIS LINE REFERENCES THE JSON CREDENTIALS FILE"""
creds = ServiceAccountCredentials.from_json_keyfile_name("C:\\Users\\Topher's PC\\Desktop\\Python\\OrderBotCreds.json",scope)
"""NEXT TWO LINES MAKE THE GOOGLE SHEET WITH CREDENTIALS EASIER TO CALL"""
client = gspread.authorize(creds)
#sheet = client.open("Alliance (Responses)").worksheet("Form_Responses_1")
################################################################################





def main (x):
    x = int(x-1)
    DChoice = [["Alliance", "Dylan", "Alliance (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Alliance_baseline.xlsx"],["Heron", "Levi Ronk", "Heron (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Heron_baseline.xlsx"],["Griffin", "Josh Raber", "Griffin (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Griffin_baseline.xlsx"],["Imperial", "Joe S.", "Imperial (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Imperial_baseline.xlsx"], ["Fabiano", "Arnnie", "Fabiano (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Fabiano_baseline.xlsx"], ["Coka-Cola", "Jason", "Coke (Responses)", "Form_Responses_1", ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/drive'], "C:\\Users\\Topher's PC\\Desktop\\Python\\Coke_baseline.xlsx"], ["PepsiCo", "Victor", "Pepsi (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Pepsi_baseline.xlsx"]]

    dis = DChoice[x][0]
    rep = DChoice[x][1]
    GOOGLE_SHEET = DChoice[x][2]
    SHEET = DChoice[x][3]
    sheet = client.open(GOOGLE_SHEET).worksheet(SHEET)
    scope = DChoice[x][4]
    filename = DChoice[x][5]

    

    """Finds the last taken inventory on Google Sheets"""
    i = 1
    while sheet.cell(i,1).value != None :
        i = i+1
    lastrow = i-1
    """Gets the headers from the inventory associated with Numbered Inventory"""
    Header = sheet.row_values(1)
    inventory = sheet.row_values(lastrow)


    """Combines the Header and Inventory into a list os lists"""
    invent = [[Header[i], inventory[i]] for i in range(len(Header))]


    ###     Notes for future adjustments   ###


    #print(invent)
    "prints entire inventory paired with inventory count"

    #print(invent[3])
    "Prints a single inventory item with its inventory count"

    #print(invent[3][0])
    "only prints the Name of the inventory item"
    #print (invent[3][1])
    "Only prints the inventory count of the inventory item"
    ###     End Notes   ###

    InventoryTaken = invent.pop(0) #Removes Time stamp
    OrderDate = invent.pop(0)      #Captures date of Inventory
    invent.sort()                  #Alphebetizes the inventory for Vendor ease

    #################################
    ### Makes Date Readable to OS ###
    #################################
    OrderDate = OrderDate.pop(1)
    OrderDate = OrderDate.split('/')
    Months = {"1":"Jan","2":"Feb","3":"Mar","4":"Apr","5":"May","6":"Jun","7":"Jul","8":"Aug","9":"Sep","10":"Oct","11":"Nov","12":"Dec"}
    OrderDate[0] = Months[OrderDate[0]]
    OrderDate = OrderDate[1] +'-'+ OrderDate[0] +'-'+ OrderDate[2]
    #################################


    ###References Baseline Microsoft Excell File###
    #file name assigned by button anf CDhoice at begining of main()
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    baseline = []
    for i in range(ws.max_row-1):
        baseline.append(ws.cell(row=int(i+2), column=2).value)
    print(len(invent)+ len(baseline))
    #####################################################################
    #####################################################################
    #   IF Over Stocked or Not needed to Order Item removed from List   #
    #####################################################################
    order = []
    for i in range(len(invent)):
        #print(str(invent[i][0])+str(int(baseline[i])-int(invent[i][1])))
        print(str(i)+"  "+str(invent[i][0])+" "+str(baseline[i]))
        c = baseline[i] - int(invent[i][1])
        order.append([invent[i][0], c])
    for i in range(len(order)):
        if (order[i][1])<0:
            order[i][1] = 0
        else: pass
    
    removal = []

    for i in range(len(order)):
        if (order[i][1]) == 0:
            r = order[i]
            removal.append(r)
        else: pass

    for i in range(len(removal)):
        order.pop(order.index(removal[i]))
    #####################################################################
    #####################################################################
    #          END REMOVING 0 ORDERS FROM MEMORY ORDERED LISTS          #
    #####################################################################


    ############ WRITING OF REPORT ###########
    Dashcount = 0
    for i in Header:
        count = len(i)  # finds length of longest item for Padding adjustment
        if count > Dashcount:
            Dashcount = count
        else:pass

    full_divider ='-'*84
    newdoc = "C:\\Users\\Topher's PC\\Desktop\\Python\\" + dis + " " + OrderDate +".txt"
    try:
        with open( newdoc,"x") as text_file:
            print(f"{dis : ^84}",file = text_file)
            print(f"{OrderDate : <15}{rep : ^50}{'CORNER GROCER' : >15}",file = text_file)
            print(full_divider, file = text_file)
            print(f"{'Product Name': ^{Dashcount}}   {'Qty' : ^3}", file = text_file)
            for i in range(len(order)):
                print(f"{order[i][0] : <{Dashcount}}  {order[i][1] : >3}", file = text_file)
            print(f"{'Order by _______________________' : >45}", file = text_file)


        text_file.close()  #Saves file

     
    except:
        root = tk.Tk()
        root.title('Tkinter Yes/No Dialog')
        root.geometry('1x1')
        MsgBox = messagebox.askquestion ('Repeat order',"Records Show an order has already been made by \n the most recent inventory data. \n Do you want to Reprint " + OrderDate + "'s Order?", icon = 'warning')
        if MsgBox == 'yes':
            os.startfile(newdoc,'print')
            root.destroy()
        else:
            messagebox.showinfo('Take Inventory and Try Again','Take Inventory and Try Again')
            root.destroy()
    os.startfile(newdoc)














def Update(x):
    
    DChoice = [["Alliance", "Dylan", "Alliance (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Alliance_baseline.xlsx"],["Heron", "Levi Ronk", "Heron (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Heron_baseline.xlsx"],["Griffin", "Josh Raber", "Griffin (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Griffin_baseline.xlsx"],["Imperial", "Joe S.", "Imperial (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Imperial_baseline.xlsx"], ["Fabiano", "Arnnie", "Fabiano (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Fabiano_baseline.xlsx"], ["Coka-Cola", "Jason", "Coke (Responses)", "Form_Responses_1", ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/drive'], "C:\\Users\\Topher's PC\\Desktop\\Python\\Coke_baseline.xlsx"], ["PepsiCo", "Victor", "Pepsi (Responses)", "Form_Responses_1", ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"], "C:\\Users\\Topher's PC\\Desktop\\Python\\Pepsi_baseline.xlsx"]]
    #for i in range(0,int(x-1)):
    i = 2
    
    dis = DChoice[i][0]
    rep = DChoice[i][1]
    GOOGLE_SHEET = DChoice[i][2]
    SHEET = DChoice[i][3]
    sheet = client.open(GOOGLE_SHEET).worksheet(SHEET)
    scope = DChoice[i][4]
    filename = DChoice[i][5]
    Header = sheet.row_values(1)







    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    Header.pop(0)
    Header.pop(0)
    Header.sort()
    Max = ws.max_row
    xc = {}

    
    for x in range(Max-1):
        r=x+1

        xc[ws.cell(row=int(r),column=1).value] = ws.cell(row=int(r),column=2).value
    del xc['Product Name']
    Dict_Header ={}
    for i in range(len(Header)):
        Dict_Header[Header[i]] = 0

    Dict_Update = Dict_Header.copy()
    Dict_Update.update(xc)

    A = list(Dict_Update.keys())
    B = list(Dict_Update.values())
    Updated = []
    
    for i in range(len(A)):
        Updated.append([A[i],B[i]])

    Updated.sort()






    
    worksheet.write(0, 0, "Product Name")
    worksheet.write(0, 1, "Baseline Inventory")
    for i in range(len(Updated)):
        worksheet.write(int(i)+1, 1, Updated[i][1])
        worksheet.write(int(i)+1, 0, Updated[i][0])
    workbook.close()
    os.startfile(filename)




############ Creates the New Window ###########
###############################################
root =Tk()                                    #
root.title("Inventory Orders")                #
root.geometry("300x300")                      #
frame = Frame(root)                           #
rightframe = Frame(root)                      #
rightframe.pack( side = RIGHT)                #
leftframe = Frame(root)                       #
leftframe.pack( side = LEFT )                 #
bottomframe = Frame(root)                     #
frame.pack( side = BOTTOM )                   #
topframe= Frame(root)                         #
frame.pack( side = TOP)                       #
text = Text(root)                             #
###############################################
    

titlebutton = Button(frame, text="Alliance", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300',  command= lambda: main(1))
titlebutton.pack()
titlebutton = Button(frame, text="Heron", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300', command= lambda: main(2))
titlebutton.pack()
titlebutton = Button(frame, text="Griffin", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300', command= lambda: main(3))
titlebutton.pack()
titlebutton = Button(frame, text="Imperial", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300', command= lambda: main(4))
titlebutton.pack()
titlebutton = Button(frame, text="Fabiano", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300', command= lambda: main(5))
titlebutton.pack()
titlebutton = Button(frame, text="Coke", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300', command= lambda: main(6))
titlebutton.pack()
titlebutton = Button(frame, text="Pepsi", fg="red",bg="black", activebackground= "black", activeforeground="red", width= '300', command= lambda: main(7))
titlebutton.pack()
titlebutton = Button(frame, text="Update", fg="green",bg="yellow", activebackground= "black", activeforeground="red", width= '300', command= lambda: Update(7))
titlebutton.pack()












































